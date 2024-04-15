# matchmaking/utils.py
import pandas as pd
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

def load_data_in_chunks(filepath, chunk_size=100):
    workbook = load_workbook(filepath, read_only=True)
    sheet = workbook.active
    data_rows = []
    current_chunk = []

    for row in sheet.iter_rows(values_only=True):
        current_chunk.append(row)
        if len(current_chunk) == chunk_size:
            data_rows.extend(current_chunk)
            current_chunk = []

    if current_chunk:
        data_rows.extend(current_chunk)

    columns = [cell.value for cell in sheet[1]]  # Assuming the first row contains headers
    dataframe = pd.DataFrame(data_rows, columns=columns)
    return dataframe

def get_professor_recommendations_for_student_name(students_df, professors_df, tfidf_vectorizer, similarity_matrix, student_name, threshold=0.5):
    if student_name in students_df['Student GUID'].values:
        student_index = students_df.index[students_df['Student GUID'] == student_name].tolist()[0]
    else:
        return {"error": "No student found with the name: {}".format(student_name)}

    sim_scores = list(enumerate(similarity_matrix[student_index]))
    sim_scores = [(prof_index, score) for prof_index, score in sim_scores if score > threshold]
    sim_scores = sorted(sim_scores, key=lambda x: x[1], reverse=True)
    professor_indices = [i[0] for i in sim_scores]
    recommended_professors = professors_df.loc[professor_indices]
    recommended_professors['Similarity_Score'] = [score for _, score in sim_scores]
    recommended_professors['Professor_GUID'] = recommended_professors['Professor GUID']
    return recommended_professors[['Professor_GUID', 'Similarity_Score']].reset_index(drop=True).to_dict(orient='records')


def prepare_tfidf(students_df, professors_df):
    tfidf_vectorizer = TfidfVectorizer(stop_words='english')
    combined_interests = pd.concat([students_df['Research Interests'], professors_df['Research Interests']], ignore_index=True)
    tfidf_vectorizer.fit(combined_interests)
    student_tfidf_matrix = tfidf_vectorizer.transform(students_df['Research Interests'])
    professor_tfidf_matrix = tfidf_vectorizer.transform(professors_df['Research Interests'])

    # Create cosine similarity matrix
    similarity_matrix = cosine_similarity(student_tfidf_matrix, professor_tfidf_matrix)

    return tfidf_vectorizer, student_tfidf_matrix, professor_tfidf_matrix, similarity_matrix

